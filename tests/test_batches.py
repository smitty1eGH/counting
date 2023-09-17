import csv
from itertools import zip_longest

from openpyxl import load_workbook 
from openpyxl.worksheet.cell_range import CellRange
import pytest

from deposit_data import cash, aba_data, counters, bag_number

cash_total=sum([k*v for k,v in cash.items()])

def set_cell_value(ws,wb,range_name,value):
    '''Extract the address from the defined names
    Break the range down to a (row,col) tuple
    Load the value into the cell
    '''
    range  = CellRange(wb.defined_names[range_name].attr_text)
    rowcol = [r for r in range.rows][0][0]
    ws.cell(rowcol[0],rowcol[1],value)


def prep_counters():
    pres=filter(lambda x:x[1]==True,counters)
    return ', '.join([x[0] for x in pres])


def check_cashrow(csv_cash_path):
    batch_total=0
    with open(csv_cash_path,'r') as data:
        for i,row in enumerate(csv.reader(data)):
            if i>0: 
                batch_total += float(row[0][1:].replace(',',''))
    assert batch_total == cash_total
  

def get_cash_block():
    SPACER=''
    out=[(SPACER,k,v,k*v) for k,v in cash.items()]
    out.append((SPACER,SPACER,'Cash total',cash_total))
    return out
    
def get_check_rows(csv_check_path):
    flds=['donor_id','check_number','net_amount']
    fixme=set()
    out=[]
    total=0.0
    with open(csv_check_path,'r') as data:
        for i,row in enumerate(csv.DictReader(data, fieldnames=flds)):
            if i>0:
                if row['donor_id'] not in aba_data:
                    fixme.add(row['donor_id'])
                total += float(row['net_amount'][1:].replace(',',''))
                out.append(( aba_data.get(row['donor_id'], f"FIX--{row['donor_id']}")
                           , row['check_number']
                           , row['net_amount']
                          ))
    out.append(('','Check Totoal', total))
    return fixme, out, total
    

@pytest.mark.skip    
def test_check_cashrow(csv_cash_path):
    check_cashrow(csv_cash_path)
    print()
    print('\n'.join([str(x) for x in get_cash_block()]))
    

def test_get_check_row(csv_cash_path, csv_check_path, deposit_template):
            
    def gen_zrows(out,cashrows):
        def maybe_row1(r1):
            if r1:
                return r1   
            else:
                return ('',)
        for row in zip_longest(out,cashrows):
            yield (*row[0], *maybe_row1(row[1]))
    
    check_cashrow(csv_cash_path)
    cashrows = get_cash_block()
    fixme, checkrows, check_total = get_check_rows(csv_check_path)
    
    wb          = load_workbook(deposit_template)
    wb.template = False
    ws          = wb.active
    set_cell_value(ws,wb,"bag_number"   ,bag_number)
    set_cell_value(ws,wb,"counted_by"   ,prep_counters())
    set_cell_value(ws,wb,"total_deposit",cash_total+check_total)
    set_cell_value(ws,wb,"item_count"   ,len(checkrows))
    for row in gen_zrows(checkrows,cashrows):
        ws.append(row)
    wb.save(f'asdf.xlsx')
    


