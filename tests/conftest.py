import os

from openpyxl import load_workbook
import pytest


@pytest.fixture
def abadata():
    return f"{os.getcwd()}\\instance\\ABA_Numbers_Donors_List.xlsx"


@pytest.fixture
def deposit_template():
    return f"{os.getcwd()}\\Deposit_Ticket_Blank.xltx"



@pytest.fixture
def ws(abadata):
    '''data_only=True means get the formula values, which we want
    '''
    wb = load_workbook(filename=abadata)
    return wb.active


@pytest.fixture
def peopledata():
    return f"{os.getcwd()}\\instance\\fbcok.csv"

@pytest.fixture
def ABA_Numbers(abadata):
    '''
    '''
    wb = load_workbook(filename=abadata)
    return wb['evaluated_ABA_Numbers']

@pytest.fixture
def csv_cash_path():
    return './instance/batch-301-cash.csv'

@pytest.fixture
def csv_check_path():
    return './instance/batch-300-checks.csv'