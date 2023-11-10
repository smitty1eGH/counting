import csv
from pathlib import Path
from itertools import zip_longest
import os

from openpyxl import load_workbook 
from openpyxl.worksheet.cell_range import CellRange

from deposit_data import bag_number, cash, aba_data, people

# CONOPS
# 0. Edit deposit_data.py to reflect
#    - counters
#    - cash received
# 1. Do regular and cash batches in PlanningCenter
# 2. Drop .csv exports regular and cash batches into instance/
# 3. Run this script to merge: 
#    3.1 cash: just summarize the amount and validate against the currency total
#    3.2 checks: do a real merge
#    3.3 (remove the cash/check .csv files)
#    3.3 deposit_data--bag_number, counters, ABA
#    3.4 Deposit_Ticket_Blank.xltx template
#    3.5 Generate output

# 0.
cash_total=sum([k*v for k,v in cash.items()])
counters  =', '.join([q[0] for q in [p for p in filter(lambda x:x[1]==True,people)]])

# 1. and 2. are done out-of-band

def set_cell_value(ws,wb,range_name,value):
    '''Extract the address from the defined names
    Break the range down to a (row,col) tuple
    Load the value into the cell
    '''
    range  = CellRange(wb.defined_names[range_name].attr_text)
    rowcol = [r for r in range.rows][0][0]
    ws.cell(rowcol[0],rowcol[1],value)

def get_check_rows(csv_check_path):
    '''This is where we pull in the checks .csv and build the 
    left column of the data
    #flds=['donor_id','check_number','net_amount']
    '''
    flds=['donor_id','donor_first_name','donor_last_name','donor_name_suffix','donor_email','check_number','amount']
    fixme=set()
    out=[]
    total=0.0
    with open(csv_check_path,'r') as data:
        for i,row in enumerate(csv.DictReader(data, fieldnames=flds)):
            if i>0:
                if row['donor_id'] not in aba_data:
                    fixme.add( row['donor_id'])
                total += float(row['amount'][1:].replace(',',''))
                out.append(( aba_data.get(row['donor_id'    ], f"FIX--{row['donor_id']}")
                           , int(         row['check_number'])
                           , float(       row['amount'      ][1:].replace(',',''))
                          ))
    out.append(('','Check Total', total))
    return fixme, out, total

def get_cash_block():
    '''Cash lines on right side of the deposit
    '''
    SPACER=''
    out=[(SPACER,k,v,k*v) for k,v in cash.items()]
    out.append((SPACER,SPACER,'Cash total',cash_total))
    return out


def validate_cashrow(csv_cash_path):
    '''Sum the batch_XXX_cash.csv data
    cash_total is at the top of the file
    
    add up the .csv and assert that the cash.items() are good
    '''
    batch_total=0
    with open(csv_cash_path,'r') as data:
        for i,row in enumerate(csv.reader(data)):
            if i>0: 
                batch_total += float(row[0][1:].replace(',',''))
    #assert batch_total == cash_total

def get_csv_paths():
    '''Locate the files in the instance/ directory.
    This is a fragile and hackish mess.
    '''
    fs = Path('./instance/').glob('**/*.csv')
    for f in fs:
      g=f.name.split('-')
      if len(g)==3:
        h=g[2].split('.')
        if h[0]=='cash':
          csv_cash_path=f
        elif h[0]=='checks':
          csv_check_path=f
        else:
          print('wut?')
    return csv_cash_path, csv_check_path

def write_deposit(csv_cash_path, csv_check_path, deposit_template):
            
    def gen_zrows(out,cashrows):
        '''This lets us merge the left and right columns
        of the output from the check- and cash lists.
        '''
        def maybe_row1(r1):
            if r1:
                return r1   
            else:
                return ('',)
        for row in zip_longest(out,cashrows):
            yield (*row[0], *maybe_row1(row[1]))
    # 3.1
    validate_cashrow(csv_cash_path)
    cashrows = get_cash_block()
    
    # 3.2
    fixme, checkrows, check_total = get_check_rows(csv_check_path)
    
    # TODO:
    # 3.3
    
    wb          = load_workbook(deposit_template)
    wb.template = False
    ws          = wb.active
    
    # 3.3
    set_cell_value(ws,wb,"bag_number"   ,bag_number)
    set_cell_value(ws,wb,"counted_by"   ,counters  )
    set_cell_value(ws,wb,"total_deposit",float(cash_total+check_total))
    set_cell_value(ws,wb,"item_count"   ,len(checkrows) -1)
    
    # 3.5
    for row in gen_zrows(checkrows,cashrows):
        ws.append(row)
        
    #TODO: real output file name
    wb.save(f'asdf.xlsx')
    print(f'{fixme=}')


def main():
    csv_cash_path, csv_check_path = get_csv_paths()
    
    # 3.4
    deposit_template = f"{os.getcwd()}\\Deposit_Ticket_Blank.xltx"
    
    write_deposit(csv_cash_path, csv_check_path, deposit_template)
    
if __name__=='__main__':
    main()